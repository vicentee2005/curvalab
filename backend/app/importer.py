"""Importación de datos tabulares (CSV / TXT / XLSX).

Para texto (CSV/TXT/DAT) se detecta automáticamente el separador y el formato
decimal (coma o punto), algo habitual en datos de laboratorio en español: se
prueban todas las combinaciones razonables y gana la que produce más celdas
numéricas. Para Excel (.xlsx) se lee con openpyxl, que sí entiende el formato
binario (un .xlsx es un ZIP con XML dentro, no texto: no se puede decodificar
como CSV). Ambas rutas devuelven columnas y filas numéricas listas para volcar
en la hoja de datos del frontend.

Sin pandas a propósito: la biblioteca pesa unos 65 MB y este módulo solo
necesitaba de ella la lectura de tablas, que aquí son pequeñas y de estructura
sencilla. Quitarla es lo que permite que el backend quepa en el límite de 250 MB
de una función serverless.
"""
from __future__ import annotations

import io

from .infer_roles import infer_roles
from .schemas import ImportResponse

# Separadores a probar. None significa "espacios en blanco" (uno o varios),
# el formato típico de los ficheros .dat de instrumentos.
_SEPARATORS: list[str | None] = [";", "\t", ",", None]
_DECIMALS = [".", ","]


class ImportError_(ValueError):
    """No se pudo interpretar el fichero."""


def _to_float(text: object, decimal: str = ".") -> float | None:
    """Convierte una celda a número, o None si no lo es."""
    if text is None:
        return None
    if isinstance(text, bool):
        return None
    if isinstance(text, (int, float)):
        return float(text)
    s = str(text).strip().strip('"').strip("'")
    if not s:
        return None
    if decimal == ",":
        s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def _looks_numeric(text: object, decimal: str = ".") -> bool:
    return _to_float(text, decimal) is not None


def _split(line: str, sep: str | None) -> list[str]:
    return line.split() if sep is None else line.split(sep)


def _build_response(
    names: list[str], columns_values: list[list[float | None]]
) -> ImportResponse:
    """Ensambla la respuesta: descarta columnas vacías, infiere roles y arma filas.

    columns_values está en orden por columnas (columns_values[c] = valores de la
    columna c). Se descartan las columnas totalmente vacías (así unos datos en la
    3ª/4ª columna se leen como X/Y y no por su posición literal) y luego se
    deducen los roles con la importación inteligente.
    """
    # Descartar columnas totalmente vacías.
    keep = [c for c in range(len(names)) if any(v is not None for v in columns_values[c])]
    if not keep:
        raise ImportError_("El fichero no contiene columnas con datos numéricos.")

    kept_names = [names[c] for c in keep]
    kept_cols = [columns_values[c] for c in keep]

    roles = infer_roles(kept_names, kept_cols)

    n_rows = max((len(c) for c in kept_cols), default=0)
    rows: list[list[float | None]] = []
    for r in range(n_rows):
        rows.append([kept_cols[c][r] if r < len(kept_cols[c]) else None
                     for c in range(len(kept_cols))])

    return ImportResponse(
        columns=kept_names,
        rows=rows,
        n_rows=n_rows,
        n_cols=len(kept_names),
        roles=roles,
    )


def _from_grid(grid: list[list[object]], decimal: str = ".") -> ImportResponse:
    """Convierte una rejilla de celdas (ya separadas) en la respuesta final.

    Si la primera fila es toda numérica, el fichero no traía cabecera y esa fila
    es un dato más: pasa a menudo con los .dat de instrumentos, que empiezan
    directamente con números.
    """
    if not grid:
        raise ImportError_("El fichero no contiene datos.")

    n_cols = max(len(row) for row in grid)
    primera = grid[0]
    hay_cabecera = not all(_looks_numeric(c, decimal) for c in primera if str(c).strip())

    if hay_cabecera:
        names = [
            str(primera[c]).strip() if c < len(primera) and primera[c] is not None else f"col{c+1}"
            for c in range(n_cols)
        ]
        cuerpo = grid[1:]
    else:
        names = [f"col{c+1}" for c in range(n_cols)]
        cuerpo = grid

    names = [n or f"col{i+1}" for i, n in enumerate(names)]

    columns_values: list[list[float | None]] = [
        [_to_float(row[c], decimal) if c < len(row) else None for row in cuerpo]
        for c in range(n_cols)
    ]
    return _build_response(names, columns_values)


def _grid_from_text(content: str, sep: str | None) -> list[list[str]]:
    """Parte el texto en celdas, saltando comentarios y líneas en blanco."""
    grid: list[list[str]] = []
    for raw in content.splitlines():
        line = raw.strip()
        if not line or line.startswith("#"):
            continue
        grid.append([c.strip() for c in _split(line, sep)])
    return grid


def _score(grid: list[list[str]], decimal: str) -> int:
    """Cuántas celdas numéricas produce esta lectura.

    Una sola columna se penaliza fuerte: casi siempre significa que el separador
    elegido no era el bueno y la fila entera se quedó pegada.
    """
    if not grid:
        return -10_000
    n_cols = max(len(row) for row in grid)
    numericas = sum(
        1 for row in grid[1:] for cell in row if _looks_numeric(cell, decimal)
    )
    return numericas - (1000 if n_cols < 2 else 0)


def parse_table(content: str) -> ImportResponse:
    """Parsea el texto de un CSV/TXT a columnas y filas numéricas.

    Prueba cada combinación de separador (';', tabulador, ',', espacios) y
    decimal (punto, coma) y se queda con la que produzca más celdas numéricas.
    """
    if not content or not content.strip():
        raise ImportError_("El fichero está vacío.")

    mejor: tuple[list[list[str]], str] | None = None
    mejor_score = -10_000
    for sep in _SEPARATORS:
        grid = _grid_from_text(content, sep)
        for decimal in _DECIMALS:
            s = _score(grid, decimal)
            if s > mejor_score:
                mejor_score = s
                mejor = (grid, decimal)

    if mejor is None or mejor_score <= -1000:
        raise ImportError_("No se pudo interpretar el fichero como tabla.")

    grid, decimal = mejor
    return _from_grid(grid, decimal)


def parse_excel(content: bytes, filename: str = "") -> ImportResponse:
    """Lee la primera hoja de un Excel .xlsx con openpyxl.

    A diferencia del CSV/TXT, un Excel es un binario (ZIP + XML) y no se puede
    decodificar como texto: hace falta un lector dedicado.
    """
    if filename.lower().endswith(".xls"):
        raise ImportError_(
            "El formato .xls antiguo no se puede leer. Abre el fichero en Excel "
            "y guárdalo como .xlsx."
        )
    try:
        from openpyxl import load_workbook

        # read_only y data_only: no interesa el formato, solo el valor calculado.
        libro = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        hoja = libro.worksheets[0]
        grid = [list(fila) for fila in hoja.iter_rows(values_only=True)]
        libro.close()
    except ImportError_:
        raise
    except Exception as exc:  # noqa: BLE001 - openpyxl lanza de todo
        raise ImportError_(f"No se pudo leer el Excel: {exc}") from exc

    # Quitar las filas completamente vacías que Excel suele dejar al final.
    grid = [fila for fila in grid if any(c is not None and str(c).strip() for c in fila)]
    if not grid:
        raise ImportError_("La hoja de Excel está vacía.")

    return _from_grid(grid)
